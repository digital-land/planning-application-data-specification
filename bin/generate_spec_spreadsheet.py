from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from planning_application_specification import Specification
from planning_application_specification.specification import (
    ResolvedComponentReference,
    ResolvedField,
    SelectionContext,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from utils import to_anchor


def format_resolved_path_name(item: ResolvedComponentReference) -> str:
    name = item.base.name
    if item.base.cardinality == "n":
        name += "[]"
    return name


def walk_resolved_component_paths(
    component_ref: str,
    package_spec: Specification,
    prefix: Tuple[List[str], List[str]],
    application_type: Optional[str] = None,
) -> List[Tuple[List[str], List[str], ResolvedField]]:
    rows: List[Tuple[List[str], List[str], ResolvedField]] = []
    selection = (
        SelectionContext(application_type=application_type) if application_type else None
    )

    for resolved_item in package_spec.resolve_container_items(
        component=component_ref,
        selection=selection,
    ):
        if not resolved_item.applies:
            continue
        if isinstance(resolved_item, ResolvedField):
            rows.append((prefix[0], prefix[1], resolved_item))
        else:
            level_prefix_names = prefix[0] + [format_resolved_path_name(resolved_item)]
            level_prefix_refs = prefix[1] + [resolved_item.base.ref]
            rows.extend(
                walk_resolved_component_paths(
                    resolved_item.component_ref,
                    package_spec,
                    (level_prefix_names, level_prefix_refs),
                    application_type,
                )
            )

    return rows


def flatten_module_to_rows_with_package(
    mod: Any,
    package_spec: Specification,
    application_type: Optional[str] = None,
) -> List[Tuple[List[str], List[str], ResolvedField]]:
    rows: List[Tuple[List[str], List[str], ResolvedField]] = []
    selection = (
        SelectionContext(application_type=application_type) if application_type else None
    )

    for item in package_spec.resolve_container_items(module=mod.ref, selection=selection):
        if not item.applies:
            continue
        if isinstance(item, ResolvedField):
            rows.append(([], [], item))
        else:
            rows.extend(
                walk_resolved_component_paths(
                    item.component_ref,
                    package_spec,
                    ([format_resolved_path_name(item)], [item.base.ref]),
                    application_type,
                )
            )

    return rows


# ---------- Excel writing ----------


def get_base_and_overrides(item: Any) -> Tuple[Any, Dict[str, Any]]:
    if isinstance(item, ResolvedField):
        return item.base, item.usage.overrides
    if hasattr(item, "original") and hasattr(item, "overrides"):
        return item.original, item.overrides
    return item, {}


def is_component_reference_instance(item: Any) -> bool:
    return hasattr(item, "component") and hasattr(item, "referenced_by_field")


def format_field_name(f: Any) -> str:
    orig, overrides = get_base_and_overrides(f)
    name = overrides.get("name") or getattr(orig, "name", "")
    cardinality = overrides.get("cardinality") or getattr(orig, "cardinality", None)
    if cardinality == "n":
        name += "[]"
    return name


def requirement_label(required: bool) -> str:
    return "MUST" if required else "MAY"


def make_row(
    application,
    app_ref_code,
    app_desc,
    top,
    top_ref,
    top_desc,
    field_chain,
    field_chain_refs,
    f: Any,
    incl_app_details: bool = True,
    incl_references: bool = False,
):
    orig, overrides = get_base_and_overrides(f)
    requirement_level = overrides.get("required", orig.required)

    row = {
        "top_level": top,
        "top_description": top_desc,
        "field_chain": field_chain,
        "description": overrides.get("description", orig.description),
        "datatype": getattr(orig, "datatype", "string"),
        "requirement": requirement_label(requirement_level),
    }

    if incl_references:
        row["top_level_ref"] = top_ref
        row["field_chain_refs"] = field_chain_refs

    if incl_app_details:
        row["application"] = application
        row["application_description"] = app_desc
        if incl_references:
            row["application_ref"] = app_ref_code

    return row


def make_empty_top_level_row(
    top: str,
    top_description: str,
    incl_app_details: bool,
    incl_references: bool,
    app_ref: str,
    app_desc: str,
) -> Dict[str, Any]:
    row = {
        "top_level": top,
        "top_description": top_description,
        "field_chain": [""],
        "description": "",
        "datatype": "",
        "requirement": "",
    }
    if incl_references:
        row["top_level_ref"] = None
        row["field_chain_refs"] = []
    if incl_app_details:
        row["application"] = app_ref
        row["application_description"] = app_desc
        if incl_references:
            row["application_ref"] = app_ref
    return row


def flatten_application_item_rows(
    item: Any,
    package_spec: Specification,
    application_type: str,
) -> List[Tuple[str, str, List[str], List[str], ResolvedField | Any]]:
    if is_component_reference_instance(item):
        ref_field = item.referenced_by_field
        top = ref_field.original.name
        top_description = (
            ref_field.overrides.get("description") or ref_field.original.description
        )
        prefix_names = [format_field_name(ref_field)]
        prefix_refs = [ref_field.original.ref]

        rows = []
        for comp_path_names, comp_path_refs, resolved_field in walk_resolved_component_paths(
            item.component.ref,
            package_spec,
            (prefix_names, prefix_refs),
            application_type,
        ):
            rows.append(
                (
                    top,
                    top_description,
                    comp_path_names + [format_field_name(resolved_field)],
                    comp_path_refs + [resolved_field.base.ref],
                    resolved_field,
                )
            )
        return rows

    if hasattr(item, "ref") and hasattr(item, "name"):
        return [
            (
                item.name,
                item.description,
                [format_field_name(item)],
                [item.ref],
                item,
            )
        ]

    return []


def append_flat_row(
    flat_rows: List[Dict[str, Any]],
    app_name: str,
    app_ref: str,
    app_desc: str,
    top: str,
    top_ref: Optional[str],
    top_desc: str,
    field_chain: List[str],
    field_chain_refs: List[str],
    row_item: Any,
    incl_app_details: bool,
    incl_references: bool,
) -> None:
    flat_rows.append(
        make_row(
            app_name,
            app_ref,
            app_desc,
            top,
            top_ref,
            top_desc,
            field_chain,
            field_chain_refs if incl_references else [],
            row_item,
            incl_app_details=incl_app_details,
            incl_references=incl_references,
        )
    )


def build_flat_rows(
    app: Any,
    incl_app_details: bool = True,
    incl_references: bool = False,
) -> List[Dict[str, Any]]:
    app_ref = app.application
    app_desc = app.description
    app_name = app.name
    package_spec = Specification.load()

    flat_rows: List[Dict[str, Any]] = []

    # 1) Application-level items (fields and possibly embedded components)
    for item in app.items:
        for top, top_desc, field_chain, field_chain_refs, row_item in (
            flatten_application_item_rows(item, package_spec, app_ref)
        ):
            top_ref = field_chain_refs[0] if incl_references and field_chain_refs else None
            append_flat_row(
                flat_rows,
                app_name,
                app_ref,
                app_desc,
                top,
                top_ref,
                top_desc,
                field_chain,
                field_chain_refs,
                row_item,
                incl_app_details,
                incl_references,
            )

    # 2) Modules: each module is a top-level block in the sheet
    for mod in app.modules:
        if not mod:
            continue

        mod_rows = flatten_module_to_rows_with_package(mod, package_spec, app_ref)

        if not mod_rows:
            row = make_empty_top_level_row(
                mod.name,
                mod.description,
                incl_app_details,
                incl_references,
                app_ref,
                app_desc,
            )
            if incl_references:
                row["top_level_ref"] = mod.ref
            flat_rows.append(row)
        else:
            for comp_path_names, comp_path_refs, f in mod_rows:
                top = mod.name
                top_ref = mod.ref if incl_references else None
                field_name = format_field_name(f)
                field_chain = comp_path_names + [field_name]
                field_chain_refs = (
                    (comp_path_refs + [f.base.ref]) if incl_references else []
                )

                append_flat_row(
                    flat_rows,
                    app_name,
                    app_ref,
                    app_desc,
                    top,
                    top_ref,
                    mod.description,
                    field_chain,
                    field_chain_refs,
                    f,
                    incl_app_details,
                    incl_references,
                )

    if not flat_rows:
        row = make_empty_top_level_row(
            "",
            "",
            incl_app_details,
            incl_references,
            app_ref or "",
            app_desc or "",
        )
        flat_rows.append(row)

    return flat_rows


def auto_width(ws):
    widths = defaultdict(int)
    for row in ws.iter_rows(values_only=True):
        for i, val in enumerate(row, start=1):
            if val is None:
                continue
            widths[i] = max(widths[i], len(str(val)))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(72, w + 2)


def format_header_row(ws):
    """
    Format the header row of the worksheet.
    ws = worksheet
    """
    thin_border = Side(border_style="thin", color="000000")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(bottom=thin_border)


def merge_cells_vertical(ws, col_num: int, row_start, row_end):
    ws.merge_cells(
        start_row=row_start, start_column=col_num, end_row=row_end, end_column=col_num
    )
    ws.cell(row=row_start, column=col_num).alignment = Alignment(
        vertical="top", wrap_text=True
    )


def set_cell_alignment(cell, alignment: Alignment):
    cell.alignment = alignment


def set_alignment_all(ws, vertical_alignment="top", start_from_row=2):
    cell_alignment = Alignment(wrap_text=True, vertical=vertical_alignment)
    for row in ws.iter_rows(min_row=start_from_row):
        for cell in row:
            set_cell_alignment(cell, cell_alignment)


def create_header_row(
    incl_app_details: bool = True,
    incl_references: bool = False,
    max_field_depth: int = 1,
):
    # In write_application_excel function, update header construction:
    header = []
    if incl_app_details:
        if incl_references:
            header += ["application", "application-ref", "application-description"]
        else:
            header += ["application", "application-description"]

    if incl_references:
        header += [
            "top-level",
            "top-level-ref",
            "top-level-description",
        ]
    else:
        header += [
            "top-level",
            "top-level-description",
        ]

    # Interleave field references and names: field1-ref, field1, field2-ref, field2, etc.
    if incl_references:
        for i in range(1, max_field_depth + 1):
            header += [f"field{i}-ref", f"field{i}"]
    else:
        header += [f"field{i}" for i in range(1, max_field_depth + 1)]

    header += ["description", "datatype", "requirement"]
    return header


def write_application_excel(
    app: Any,
    out_dir: Path,
    incl_app_details: bool = True,
    incl_references: bool = False,
) -> Path:
    """
    Writes one XLSX with columns:
      application, application-description, top-level, field1..N, description, datatype, requirement
    Merges:
      - application/application-description across all rows
      - top-level across its contiguous block
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Specification"

    app_ref = app.application
    app_desc = app.description
    app_name = app.name

    flat_rows = build_flat_rows(
        app,
        incl_app_details=incl_app_details,
        incl_references=incl_references,
    )

    # Figure out max depth of field_chain to create field1..N columns
    max_depth = max((len(r["field_chain"]) for r in flat_rows), default=1)

    # Header
    ws.append(
        create_header_row(
            incl_app_details=incl_app_details,
            incl_references=incl_references,
            max_field_depth=max_depth,
        )
    )

    # format the heeader row
    format_header_row(ws)

    # Write rows while remembering blocks for merging
    start_row_for_app = ws.max_row + 1
    # We also merge top-level blocks: whenever top_level value changes, close the previous block
    current_top = None
    top_block_start = None
    module_cols = [1, 2]
    if incl_app_details:
        module_cols = [3, 4]

    def close_top_block(end_row: int, cols: List[int]):
        if top_block_start is not None and end_row >= top_block_start:
            # merge the top-level and top-level-description cols
            for col in cols:
                merge_cells_vertical(ws, col, top_block_start, end_row)

    for row in flat_rows:
        # flush/rotate top-level block if the value changes
        if row["top_level"] != current_top:
            # close previous block
            prev_end = ws.max_row
            close_top_block(prev_end, cols=module_cols)
            # start new block
            current_top = row["top_level"]
            top_block_start = prev_end + 1

        # Expand field_chain into field columns with padding
        chain = row["field_chain"]
        padded_names = chain + [""] * (max_depth - len(chain))

        spreadsheet_row = []
        if incl_app_details:
            spreadsheet_row.append(row["application"])
            if incl_references:
                spreadsheet_row.append(row["application_ref"])
            spreadsheet_row.append(row["application_description"])

        spreadsheet_row.append(row["top_level"])
        if incl_references:
            spreadsheet_row.append(row["top_level_ref"])
        spreadsheet_row.append(row["top_description"])

        # Interleave field references and names
        if incl_references:
            chain_refs = row.get("field_chain_refs", [])
            padded_refs = chain_refs + [""] * (max_depth - len(chain_refs))

            # Interleave: ref1, name1, ref2, name2, etc.
            for i in range(max_depth):
                spreadsheet_row.append(padded_refs[i])  # field{i}-ref
                spreadsheet_row.append(padded_names[i])  # field{i}
        else:
            # Just add field name columns
            spreadsheet_row.extend(padded_names)

        spreadsheet_row.extend(
            [row["description"], row["datatype"], row["requirement"]]
        )

        ws.append(spreadsheet_row)

    # Close last top-level block
    close_top_block(ws.max_row, cols=module_cols)

    # Merge application/application-description across all rows
    if incl_app_details:
        end_row_for_app = ws.max_row
        if end_row_for_app >= start_row_for_app:
            # merge application cells
            merge_cells_vertical(ws, 1, start_row_for_app, end_row_for_app)
            merge_cells_vertical(ws, 2, start_row_for_app, end_row_for_app)

    # Excel formatting
    set_alignment_all(ws)
    auto_width(ws)

    filename = f"{app_name} ({app_ref})"
    if incl_references:
        filename += "--verbose"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{to_anchor(filename)}.xlsx"
    wb.save(out_path)
    return out_path


# ---------- Run for all applications ----------
if __name__ == "__main__":
    from loader import load_specification_model

    model = load_specification_model()
    output_dir = Path("generated/spreadsheet")
    for app in model["applications"].values():
        # check it isn't a sub-type
        if app.extends is None:
            # Generate standard version (names only)
            path = write_application_excel(app, output_dir, incl_app_details=False)
            print("Wrote:", path)

            # Generate version with references
            path2 = write_application_excel(
                app,
                output_dir / "verbose",
                incl_app_details=False,
                incl_references=True,
            )
            print("Wrote references version:", path2)
