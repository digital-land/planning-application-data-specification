from pathlib import Path

from openpyxl import load_workbook

from generate_spec_spreadsheet import write_application_excel
from loader import load_specification_model


def _load_sheet_rows(path: Path):
    workbook = load_workbook(path)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    merged_ranges = sorted(str(range_ref) for range_ref in worksheet.merged_cells.ranges)
    return rows, merged_ranges


def test_write_application_excel_preserves_filtered_rows_and_merge_ranges(project_root, tmp_path):
    model = load_specification_model()
    app = model["applications"]["full"]

    workbook_path = write_application_excel(
        app,
        tmp_path,
        incl_app_details=False,
        incl_references=False,
    )

    rows, merged_ranges = _load_sheet_rows(workbook_path)

    assert rows[0] == (
        "top-level",
        "top-level-description",
        "field1",
        "field2",
        "field3",
        "field4",
        "description",
        "datatype",
        "requirement",
    )
    assert rows[1] == (
        "Application",
        "The details of the application payload to be submitted",
        "Application",
        "Reference",
        None,
        None,
        "A unique reference for the data item",
        "string",
        "MUST",
    )
    assert (
        "Description of the proposal",
        "What development, works or change of use is proposed",
        "Proposal description",
        None,
        None,
        None,
        "A description of what is being proposed, including the development, works, or change of use",
        "string",
        "MUST",
    ) in rows
    assert not any(row[2] == "Related application[]" for row in rows if len(row) > 2)
    assert "A159:A165" in merged_ranges
    assert "B159:B165" in merged_ranges


def test_write_application_excel_verbose_includes_reference_columns(tmp_path):
    model = load_specification_model()
    app = model["applications"]["full"]

    workbook_path = write_application_excel(
        app,
        tmp_path,
        incl_app_details=False,
        incl_references=True,
    )

    rows, _ = _load_sheet_rows(workbook_path)

    assert rows[0] == (
        "top-level",
        "top-level-ref",
        "top-level-description",
        "field1-ref",
        "field1",
        "field2-ref",
        "field2",
        "field3-ref",
        "field3",
        "field4-ref",
        "field4",
        "description",
        "datatype",
        "requirement",
    )
    assert rows[1] == (
        "Application",
        "application",
        "The details of the application payload to be submitted",
        "application",
        "Application",
        "reference",
        "Reference",
        None,
        None,
        None,
        None,
        "A unique reference for the data item",
        "string",
        "MUST",
    )


def test_write_application_excel_preserves_nested_component_rows(tmp_path):
    model = load_specification_model()
    app = model["applications"]["notice-trees-in-con-area"]

    workbook_path = write_application_excel(
        app,
        tmp_path,
        incl_app_details=False,
        incl_references=False,
    )

    rows, _ = _load_sheet_rows(workbook_path)

    assert (
        None,
        None,
        "Tree details[]",
        "Species",
        None,
        None,
        "The species of the tree",
        "string",
        "MAY",
    ) in rows
    assert not any(row[3] == "Reason" for row in rows if len(row) > 3 and row[0] == "Identification of tree(s) and description of works")
